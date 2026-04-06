#!/usr/bin/env python3
"""
Scan the database for tracks missing mood, BPM, or acoustic ID metadata.
Outputs an XLSX with three sheets listing the affected directories.

Usage:
    python3 scripts/missing_metadata_report.py
    python3 scripts/missing_metadata_report.py --output /path/to/report.xlsx

Requires: openpyxl, psycopg2 (or psycopg2-binary)
"""

import argparse
import os
import sys

try:
    import psycopg2
except ImportError:
    sys.exit("psycopg2 not installed. Run: pip3 install --user --break-system-packages psycopg2-binary")

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
except ImportError:
    sys.exit("openpyxl not installed. Run: pip3 install --user --break-system-packages openpyxl")


# Metadata key groups to check (standard + iTunes variants)
MOOD_KEYS = ["MOOD_HAPPY", "----:com.apple.iTunes:MOOD_HAPPY"]
BPM_KEYS = ["IntegerBpm", "BPM", "Bpm", "FBPM", "fBPM", "----:com.apple.iTunes:fBPM", "fBPM2"]
ACOUSTID_KEYS = ["acoustid_id", "Acoustid Id", "ACOUSTID_ID", "Acoustid Fingerprint", "ACOUSTID_FINGERPRINT"]


def build_missing_query(keys: list[str]) -> str:
    """Build a SQL condition that checks none of the given keys exist in metadata."""
    conditions = " OR ".join(f"""metadata ? '{k}'""" for k in keys)
    return f"""
        SELECT DISTINCT regexp_replace("filePath", '/[^/]+$', '') AS dir
        FROM "LocalReleaseTrack"
        WHERE "filePath" IS NOT NULL
          AND NOT ({conditions})
        ORDER BY dir
    """


def get_database_url() -> str:
    """Read DATABASE_URL from environment or web/.env file."""
    url = os.environ.get("DATABASE_URL")
    if url:
        return url

    # Try loading from web/.env
    env_path = os.path.join(os.path.dirname(__file__), "..", "web", ".env")
    if os.path.exists(env_path):
        with open(env_path) as f:
            for line in f:
                line = line.strip()
                if line.startswith("DATABASE_URL="):
                    return line.split("=", 1)[1].strip().strip('"').strip("'")

    sys.exit("DATABASE_URL not found in environment or web/.env")


def main():
    parser = argparse.ArgumentParser(description="Generate missing metadata report (XLSX)")
    parser.add_argument("--output", "-o", default="missing_metadata.xlsx", help="Output file path")
    args = parser.parse_args()

    db_url = get_database_url()
    conn = psycopg2.connect(db_url)
    cur = conn.cursor()

    sheets = [
        ("Mood", MOOD_KEYS),
        ("BPM", BPM_KEYS),
        ("Acoustic ID", ACOUSTID_KEYS),
    ]

    wb = Workbook()
    header_font = Font(bold=True, size=12)
    header_fill = PatternFill(start_color="F59E0B", end_color="F59E0B", fill_type="solid")

    for i, (name, keys) in enumerate(sheets):
        ws = wb.active if i == 0 else wb.create_sheet()
        ws.title = name

        ws["A1"] = "Directory"
        ws["A1"].font = header_font
        ws["A1"].fill = header_fill
        ws.column_dimensions["A"].width = 80

        query = build_missing_query(keys)
        cur.execute(query)
        rows = cur.fetchall()

        for row_idx, (directory,) in enumerate(rows, start=2):
            ws.cell(row=row_idx, column=1, value=directory)

        print(f"{name}: {len(rows)} directories")

    cur.close()
    conn.close()

    wb.save(args.output)
    print(f"\nSaved to {args.output}")


if __name__ == "__main__":
    main()
